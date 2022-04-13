from bs4 import BeautifulSoup
import requests
from openpyxl import load_workbook
import shutil
import random


def excel(terms, time):
    shutil.copyfile('KahootQuizTemplate.xlsx',
                    '/Users/tylerbeer/Downloads/Qahoot.xlsx')
    wb = load_workbook('/Users/tylerbeer/Downloads/Qahoot.xlsx')
    ws = wb.active
    ts = enumerate(terms)
    for i, pair in ts:
        choices = random.sample(list(terms.values()), 3)
        choices.append(terms[pair])
        random.shuffle(choices)
        ws[f'B{9+i}'].value = pair
        ws[f'C{9+i}'].value = choices[0]
        ws[f'D{9+i}'].value = choices[1]
        ws[f'E{9+i}'].value = choices[2]
        ws[f'F{9+i}'].value = choices[3]
        ws[f'G{9+i}'].value = time
        ws[f'H{9+i}'].value = choices.index(terms[pair]) + 1

    wb.save("/Users/tylerbeer/Downloads/Qahoot.xlsx")
    wb.close()


def main():
    url = input("Enter quizlet link:")
    t = int(input("How much time do you want per question (in seconds): "))

    headers = {
        'User-Agent': 'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:80.0) Gecko/20100101 Firefox/80.0'}
    soup = BeautifulSoup(requests.get(
        url, headers=headers).content, 'html.parser')

    d = {}
    q = ""
    while True:
        q = input(
            "Type '1' to have the term as the question, or '2' for the definition as the question: ")
        if q == '1':
            for _, (question, answer) in enumerate(zip(soup.select('a.SetPageTerm-wordText'), soup.select('a.SetPageTerm-definitionText')), 1):
                d[question.get_text(strip=True, separator='\n')] = answer.get_text(
                    strip=True, separator='\n')
            break
        elif q == '2':
            for _, (question, answer) in enumerate(zip(soup.select('a.SetPageTerm-wordText'), soup.select('a.SetPageTerm-definitionText')), 1):
                d[answer.get_text(strip=True, separator='\n')] = question.get_text(
                    strip=True, separator='\n')
            break
        else:
            print("Try again.")

    excel(d, t)


if __name__ == "__main__":
    main()
